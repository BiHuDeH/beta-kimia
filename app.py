import os
import pandas as pd
import streamlit as st
from io import BytesIO
from datetime import datetime
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from cryptography.fernet import Fernet
import time

# Version and Update Information
SCRIPT_VERSION = "v1.2"
UPDATE_DATE = "2024-11-01"

# Generate a key for encryption
encryption_key = Fernet.generate_key()
cipher = Fernet(encryption_key)

# Custom CSS for UI and Branding
def set_custom_style():
    st.markdown(
        f"""
        <style>
            /* Page title styling */
            .report-title {{
                color: #0072C6; /* Light blue */
                font-size: 1.75em;
                font-weight: bold;
                text-align: center;
                margin-bottom: -10px;
            }}
            .report-subtitle {{
                color: #003366; /* Dark blue */
                font-size: 11px;
                text-align: center;
            }}
            /* File uploader styling */
            .file-uploader .upload-text {{
                color: #B22222; /* Red color */
                font-weight: bold;
                font-size: 13px;
            }}
            .file-uploader .limit-text {{
                color: gray;
                font-size: 11px;
            }}
            .file-uploader .accepted-file {{
                color: gray;
                font-size: 11px;
                font-weight: bold;
            }}
            /* Button and progress bar styling */
            .stButton > button {{
                font-size: 14px;
                padding: 10px 20px;
                border-radius: 8px;
            }}
            .stButton > button:hover {{
                opacity: 0.8;
            }}
            .progress-bar {{
                background-color: #d3d3d3;
                border-radius: 5px;
                height: 10px;
                width: 100%;
            }}
            .progress-bar-fill {{
                background-color: #5cb85c; /* Green progress bar */
                height: 100%;
                width: 0;
                border-radius: 5px;
                transition: width 0.4s ease;
            }}
            /* Script version and last update info */
            .version-info {{
                font-size: 9px;
                color: #555;
                text-align: right;
                margin-top: -20px;
            }}
        </style>
        """,
        unsafe_allow_html=True
    )

# Page title and subtitle
def display_title():
    st.markdown("<div class='report-title'>Roberto Weekly Financial Data Report</div>", unsafe_allow_html=True)
    st.markdown("<div class='report-subtitle'>by Kimia Nourian</div>", unsafe_allow_html=True)

def display_version_info():
    st.markdown(f"<div class='version-info'>Script Version: {SCRIPT_VERSION} | Last Update: {UPDATE_DATE}</div>", unsafe_allow_html=True)

def encrypt_file(file_data):
    """Encrypt the uploaded file."""
    return cipher.encrypt(file_data)

def decrypt_file(encrypted_data):
    """Decrypt the encrypted file in memory."""
    return cipher.decrypt(encrypted_data)

# Function to display a progress bar during file upload
def show_progress_bar():
    st.markdown("<div class='progress-bar'><div class='progress-bar-fill' style='width: 0%;'></div></div>", unsafe_allow_html=True)
    for percent in range(1, 101, 10):
        time.sleep(0.1)
        st.markdown(f"<div class='progress-bar'><div class='progress-bar-fill' style='width: {percent}%;'></div></div>", unsafe_allow_html=True)

# Customized file uploader with new text and button labels
def custom_file_uploader():
    st.markdown("<div class='file-uploader'><p class='upload-text'>Upload your Data file by simply dropping it here or using the Upload Button.</p></div>", unsafe_allow_html=True)
    uploaded_file = st.file_uploader("Upload File", type=["xlsx", "pdf"])
    st.markdown("<p class='limit-text'>Limit 200MB</p>", unsafe_allow_html=True)
    st.markdown("<p class='accepted-file'>Excel & PDF files accepted</p>", unsafe_allow_html=True)
    return uploaded_file

# Custom report generation with styling
def create_styled_report(df):
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    
    # Style variables
    header_font = Font(bold=True, size=14)
    regular_font = Font(size=12)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style="thin", color="333333"),
                         right=Side(style="thin", color="333333"),
                         top=Side(style="thin", color="333333"),
                         bottom=Side(style="thin", color="333333"))
    thick_border = Border(left=Side(style="thick", color="333333"),
                          right=Side(style="thick", color="333333"),
                          top=Side(style="thick", color="333333"),
                          bottom=Side(style="thick", color="333333"))

    # Set column headers and arrange them based on the specified order
    ordered_columns = ['تاریخ', 'کارت به کارت', 'فروش', 'مالیات', 'کارمزد', 'برداشت روز', 'مانده آخر روز', 'واریزی اسنپ']
    df = df[ordered_columns]

    # Populate table with headers and data, setting column headers first with styling
    ws.append(ordered_columns)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thick_border
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

    # Populate data rows with formatting
    for row in df.itertuples(index=False, name=None):
        ws.append(row)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = regular_font
            cell.alignment = center_align
            cell.border = thin_border
            if cell.column == 1:
                cell.number_format = '[$-fa-IR,700]yyyy/mm/dd;@'
            else:
                cell.number_format = '#,##0.00'

    # Set column widths and row heights
    for col in ws.columns:
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = 20
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 30

    # Define the table range and add a table with Total Row
    last_column_letter = get_column_letter(ws.max_column)
    table_ref = f"A1:{last_column_letter}{ws.max_row}"
    tab = Table(displayName="ReportTable", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    tab.showTotals = True

    # Set Total Row sum options for numeric columns
    for idx, col in enumerate(ordered_columns[1:], start=1):
        tab.tableColumns[idx].totalsRowFunction = "sum"

    ws.add_table(tab)

    wb.save(output)
    output.seek(0)
    return output.getvalue()

# Main app setup
def main():
    st.set_page_config(page_title="Roberto Weekly Financial Data Report", page_icon="📊")
    set_custom_style()
    display_title()
    display_version_info()

    uploaded_file = custom_file_uploader()
    
    if uploaded_file:
        show_progress_bar()
        encrypted_data = encrypt_file(uploaded_file.getvalue())
        decrypted_data = decrypt_file(encrypted_data)

        if uploaded_file.type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            try:
                df = pd.read_excel(BytesIO(decrypted_data), skiprows=2)
                expected_columns = ['Index', 'Branch Code', 'Branch', 'Date', 'Time', 'Document Number', 
                                    'Receipt Number', 'Check Number', 'Description', 'Withdrawal', 
                                    'Deposit', 'Balance', 'Notes']
                
                if len(df.columns) == len(expected_columns):
                    df.columns = expected_columns
                else:
                    st.error("Uploaded file does not match the expected column structure.")
                    return

            except Exception as e:
                st.error(f"Error reading the Excel file: {e}")
                return
        elif uploaded_file.type == "application/pdf":
            df = extract_data_from_pdf(BytesIO(decrypted_data))

        report = process_data(df)
        excel_data = create_styled_report(report)

        col1, col2 = st.columns(2)
        with col1:
            preview_button = st.button("Preview Report", key="preview", help="Preview the generated report")
            if preview_button:
                st.write("### Report Preview")
                st.dataframe(report)

        with col2:
            download_button = st.download_button(
                label="Download Report as Excel",
                data=excel_data,
                file_name="Financial_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

if __name__ == "__main__":
    main()
